import tkinter as tk
from tkinter import filedialog
import openpyxl

def select_file_a():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    label_file_a.config(text=file_path)
    global wb_a
    wb_a = openpyxl.load_workbook(file_path)

def select_file_b():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    label_file_b.config(text=file_path)
    global wb_b
    wb_b = openpyxl.load_workbook(file_path)

def compare_files():
    ws_output = wb_a.active
    order_quantity_a = None
    order_quantity_b = None
    for sheet_a, sheet_b in zip(wb_a, wb_b):
        for row_a, row_b in zip(sheet_a.iter_rows(), sheet_b.iter_rows()):
            if row_a[0].row == 1:
                for cell in row_a:
                    if cell.value == "Order Quantity":
                        order_quantity_a = cell.column
                for cell in row_b:
                    if cell.value == "Order Quantity":
                        order_quantity_b = cell.column
            else:
                if order_quantity_a and order_quantity_b:
                    if row_a[order_quantity_a - 1].value != row_b[order_quantity_b - 1].value:
                        ws_output[row_a[order_quantity_a - 1].coordinate].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                else:
                    for cell_a, cell_b in zip(row_a, row_b):
                        if cell_a.value != cell_b.value:
                            ws_output[cell_a.coordinate].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    total_cells = 0
    mismatch_cells = 0

    for sheet_a, sheet_b in zip(wb_a, wb_b):
        for row_a, row_b in zip(sheet_a.iter_rows(), sheet_b.iter_rows()):
            for cell_a, cell_b in zip(row_a, row_b):
                total_cells += 1
                if order_quantity_a and order_quantity_b:
                    if row_a[order_quantity_a - 1].value != row_b[order_quantity_b - 1].value:
                        mismatch_cells += 1
                else:
                    if cell_a.value != cell_b.value:
                        mismatch_cells += 1

    similarity_percent = (total_cells - mismatch_cells) / total_cells * 100

    wb_a.save('output.xlsx')
    wb_a.close()
    wb_b.close()

    label_similarity.config(text=f"Similarity: {similarity_percent:.2f}%")
    
    if not order_quantity_a or not order_quantity_b:
        label_warning.config(text="Column 'Order Quantity' not found in one or both files")
    else:
        label_warning.config(text="")

root = tk.Tk()

label_file_a = tk.Label(root, text="Select file A")
label_file_a.pack()

button_file_a = tk.Button(root, text="Select file A", command=select_file_a)
button_file_a.pack()

label_file_b = tk.Label(root, text="Select file B")
label_file_b.pack()

button_file_b = tk.Button(root, text="Select file B", command=select_file_b)
button_file_b.pack()

button_compare = tk.Button(root, text="Compare files", command=compare_files)
button_compare.pack()

label_similarity = tk.Label(root, text="")
label_similarity.pack()

label_warning = tk.Label(root, text="")
label_warning.pack()

root.mainloop()