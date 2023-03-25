import openpyxl

wb_a = openpyxl.load_workbook('a.xlsx')
wb_b = openpyxl.load_workbook('b.xlsx')

ws_output = wb_a.active
for sheet_a, sheet_b in zip(wb_a, wb_b):
    for row_a, row_b in zip(sheet_a.iter_rows(), sheet_b.iter_rows()):
        for cell_a, cell_b in zip(row_a, row_b):
            if cell_a.value != cell_b.value:
                ws_output[cell_a.coordinate].fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
#look a like percent %
total_cells = 0
mismatch_cells = 0

for sheet_a, sheet_b in zip(wb_a, wb_b):
    for row_a, row_b in zip(sheet_a.iter_rows(), sheet_b.iter_rows()):
        for cell_a, cell_b in zip(row_a, row_b):
            total_cells += 1
            if cell_a.value != cell_b.value:
                mismatch_cells += 1
                
similarity_percent = (total_cells - mismatch_cells) / total_cells * 100

#saving the output file
wb_a.save('output.xlsx')
wb_a.close()
wb_b.close()

print(f"The similarity between the documents is: {similarity_percent:.2f}%")