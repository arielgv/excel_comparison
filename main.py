import openpyxl
from pdfminer.high_level import extract_text
from difflib import SequenceMatcher

# Function to extract text from PDF file
def extract_text_from_pdf(file_path):
    with open(file_path, 'rb') as f:
        text = extract_text(f)
    return text

# Function to compare two strings using SequenceMatcher
def compare_strings(s1, s2):
    matcher = SequenceMatcher(None, s1, s2)
    return matcher.ratio()

# Function to compare two Excel files
def compare_excel_files(file1, file2):
    wb1 = openpyxl.load_workbook(file1)
    wb2 = openpyxl.load_workbook(file2)

    sheet1 = wb1.active
    sheet2 = wb2.active

    max_row = max(sheet1.max_row, sheet2.max_row)
    max_col = max(sheet1.max_column, sheet2.max_column)

    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            cell1 = sheet1.cell(row=i, column=j).value
            cell2 = sheet2.cell(row=i, column=j).value

            if cell1 != cell2:
                similarity = compare_strings(str(cell1), str(cell2))
                if similarity < 0.8:
                    sheet1.cell(row=i, column=j).fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    wb1.save(file1)

# Function to compare an Excel file and a PDF file
def compare_excel_pdf(excel_file, pdf_file):
    text = extract_text_from_pdf(pdf_file)
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2):
        part_number = row[0].value
        quantity = row[1].value
        if part_number is None or quantity is None:
            continue
        if str(part_number) not in text or str(quantity) not in text:
            row[0].fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            row[1].fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    wb.save(excel_file)

# Main function to run the program
def main():
    # Get input file paths from user
    excel_file = input("Enter the path of the Excel file: ")
    pdf_file = input("Enter the path of the PDF file: ")

    # Check if the input files are valid
    if not excel_file.endswith('.xlsx'):
        print("Invalid Excel file")
        return
    #if not pdf_file.endswith('.pdf'):
    #    print("Invalid PDF file")
    #    return

    # Compare the files
    compare_excel_pdf(excel_file, pdf_file)
    compare_excel_files(excel_file, excel_file)

    # Calculate percentual of 'looks a like'
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active

    total_rows = sheet.max_row - 1
    highlighted_rows = 0
    for row in sheet.iter_rows(min_row=2):
        part_number = row[0].value
        quantity = row[1].value
        fill = row[0].fill.start_color.rgb
        if fill == 'FFFF00':
            highlighted_rows += 1

    looks_a_like_percent = round((1 - (highlighted_rows / total_rows)) * 100, 2)
    print(f"Looks a like: {looks_a_like_percent}%")

    # Save the output Excel file
    output_file = 'output.xlsx'
    wb.save(output_file)

    # Generate web-based report
    report = "<table><tr><th>Part Number</th><th>Quantity</th></tr>"
    for row in sheet.iter_rows(min_row=2):
        part_number = row[0].value
        quantity = row[1].value
        fill = row[0].fill.start_color.rgb
        if fill == 'FFFF00':
            report += f"<tr><td>{part_number}</td><td>{quantity}</td></tr>"
    report += "</table>"

    # Print the web-based report
    print(report)

if __name__ == "__main__":
    main()